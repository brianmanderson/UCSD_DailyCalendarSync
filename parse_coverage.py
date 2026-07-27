"""Parse a weekly coverage workbook and extract one person's assignments.

Expected layout (one tab per week):
  - Row 1, columns C-I hold the seven dates (Mon-Sun); C1 is that week's Monday.
    Tabs are looked up by the C1 date, not by tab name.
  - Row 2, columns C-I hold the weekday names.
  - Column A holds section headers. Everything at/below the section whose
    header starts with "Resident" is ignored.
  - Column B holds the coverage task name; cells C-I in that row hold the
    initials of the assigned person.
  - Columns K and beyond (legend/notes) are ignored.
"""

import datetime
import sys

import openpyxl

DAY_COLUMNS = range(3, 10)  # columns C..I
FIRST_TASK_ROW = 3  # rows 1-2 are the date and weekday headers


def _as_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


WEEKS = 3  # this week plus the next two


def _label(offset):
    return {0: "this week", 1: "next week"}.get(offset, f"in {offset} weeks")


def parse_assignments(xlsx_path, initials, today, weeks=WEEKS):
    """Return this week's and the following weeks' assignments for the given
    initials — `weeks` tabs in total, starting with the one containing today.

    Each week also carries the days it covers and every task name listed on
    that tab (whoever they are assigned to), so the caller can tell which
    calendar events came from the sheet and which are unrelated.
    """
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    initials = initials.strip().upper()
    monday = today - datetime.timedelta(days=today.weekday())
    parsed = [
        _parse_week(workbook, monday + datetime.timedelta(days=7 * offset), _label(offset), initials)
        for offset in range(weeks)
    ]
    return {"today": today.isoformat(), "initials": initials, "weeks": parsed}


def _parse_week(workbook, monday, label, initials):
    week = {
        "label": label,
        "monday": monday.isoformat(),
        "sheet": None,
        "days": [],
        "tasks": [],
        "assignments": [],
    }
    sheet = None
    for name in workbook.sheetnames:
        if _as_date(workbook[name].cell(row=1, column=3).value) == monday:
            sheet = workbook[name]
            break
    if sheet is None:
        week["error"] = f"no tab whose first date cell (C1) is {monday.isoformat()}"
        return week
    week["sheet"] = sheet.title

    for col in DAY_COLUMNS:
        date = _as_date(sheet.cell(row=1, column=col).value)
        if date is None:
            continue
        week["days"].append(
            {
                "date": date.isoformat(),
                "day": str(sheet.cell(row=2, column=col).value or "").strip(),
            }
        )

    section = None
    for row in range(FIRST_TASK_ROW, sheet.max_row + 1):
        header = sheet.cell(row=row, column=1).value
        if header and str(header).strip():
            section = str(header).strip()
            if section.lower().startswith("resident"):
                break
        task = sheet.cell(row=row, column=2).value
        if not (task and str(task).strip()):
            continue
        if str(task).strip() not in week["tasks"]:
            week["tasks"].append(str(task).strip())
        for col in DAY_COLUMNS:
            value = sheet.cell(row=row, column=col).value
            if value is None or str(value).strip().upper() != initials:
                continue
            date = _as_date(sheet.cell(row=1, column=col).value)
            if date is None:
                continue
            week["assignments"].append(
                {
                    "task": str(task).strip(),
                    "date": date.isoformat(),
                    "day": str(sheet.cell(row=2, column=col).value or "").strip(),
                    "section": section,
                }
            )
    return week


if __name__ == "__main__":
    import json

    if len(sys.argv) < 3:
        sys.exit("usage: python parse_coverage.py <workbook.xlsx> <initials> [YYYY-MM-DD]")
    when = (
        datetime.date.fromisoformat(sys.argv[3])
        if len(sys.argv) > 3
        else datetime.date.today()
    )
    print(json.dumps(parse_assignments(sys.argv[1], sys.argv[2], when), indent=2))
